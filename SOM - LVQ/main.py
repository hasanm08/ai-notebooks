from datapoint import DataPoint
from som import SOM
from lvq import LVQ
# import xlrd
import pandas as pd
import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook.workbook import Workbook


def load_data(filename):
    wb = Workbook()
    book = wb.active
    print('[1] Loading Data ....')
    # book = xlrd.open_workbook(filename)
    df = pd.read_excel(filename)
    print(df)
    for r in dataframe_to_rows(df, index=True, header=True):
        book.append(r)
    input_sheet = book.sheet_by_name("input")
    output_sheet = book.sheet_by_name("A_F")
    num_rows = input_sheet.nrows
    num_cols = input_sheet.ncols
    col_values = []
    for col in range(1, num_cols):
        col_values.append(input_sheet.col_values(col)[1:])
    x = np.array(col_values, dtype='|S4')
    y = x.astype(np.float)
    maxs = [max(y[col]) for col in range(0, num_cols-1)]
    mins = [min(y[col]) for col in range(0, num_cols-1)]
    data_points = []
    for row in range(1, num_rows):
        values = []
        for col in range(1, num_cols):
            values.append((float(input_sheet.cell(row, col).value) -
                          mins[col-1])/(maxs[col-1]-mins[col-1]))
        d = DataPoint(values, int(output_sheet.cell(row, 0).value))
        data_points.append(d)
    print(num_rows-1, " points with dimesion=", num_cols-1, " are added")
    return data_points


data_points = load_data("722.xlsx")
s = SOM(2, 8, 27)
s.load_input_data(data_points)
print(s.data)
s.fit(2, 0.1)

v = LVQ(2, 6, 5)
v.load_data(data_points)
print(v.data)
v.train(5, 0.01, 2)

s.predict(data_points)
v.predict(data_points)
